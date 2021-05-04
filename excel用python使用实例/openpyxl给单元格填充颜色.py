#!/usr/bin/python3
# coding: utf-8

# 1-加载库文件
from openpyxl import Workbook
from openpyxl.styles import  PatternFill

#2-新建一个工作簿
wb = Workbook()
ws = wb.active

#随便赋个值
d4 = ws['D4']
d4.value = '43'

# 也可以通过下标（第五行第二列）往“B5”单元格写入数据；
ws.cell(row=5, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#3-设置样式，并且加载到对应单元格
fill = PatternFill("solid", fgColor="1874CD")
d4.fill = fill

#保存文件
wb.save('test.xlsx')

# fill_type 有如下的方式 一般纯色填充使用 solid 其他样式自行尝试
# {'lightGrid', 'gray0625', 'lightTrellis', 'lightDown', 'lightVertical', 'darkTrellis', 'darkHorizontal', 'darkVertical', 'darkGrid', 'darkGray', 'solid', 'darkUp', 'lightGray', 'mediumGray', 'darkDown', 'lightHorizontal', 'lightUp', 'gray125'}

# fgColor 是对应的颜色代码：http://www.114la.com/other/rgb.htm 可以到该网站查


def main():
    pass


if __name__ == '__main__':
    main()
