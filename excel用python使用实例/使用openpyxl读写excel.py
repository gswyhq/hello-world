#!/usr/bin/python3
# coding: utf-8

# xlrd和xlwt处理的是xls文件，单个sheet最大行数是65535，如果有更大需要的，建议使用openpyxl函数，最大行数达到1048576。
# 如果数据量超过65535就会遇到：ValueError: row index was 65536, not allowed by .xls format

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

# 与xlrd, xlwt不同的是，openpyxl的读写excel ,下标是从1开始的。

def readExel():
    filename = r'D:\test.xlsx'
    inwb = openpyxl.load_workbook(filename)  # 读文件
    sheetnames = inwb.get_sheet_names()  # 获取读文件中所有的sheet，通过名字的方式
    ws = inwb.get_sheet_by_name(sheetnames[0])  # 获取第一个sheet内容

    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    for r in range(1,rows):
        for c in range(1,cols):
            print(ws.cell(r,c).value)
        if r==10:
            break

def writeExcel():
    outwb = openpyxl.Workbook()  # 打开一个将写的文件
    outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet
    for row in range(1,70000):
        for col in range(1,4):
            outws.cell(row, col).value = ILLEGAL_CHARACTERS_RE.search("", "{}".format(row*2))  # 写文件
        print(row)
    saveExcel = "D:\\test2.xlsx"
    outwb.save(saveExcel)  # 一定要记得保存

def write_excel(datas, save_file = r'D:\Users\excel文件名.xlsx', title = "工作表1"):
    """将二维列表数据写入excel"""
    wb = Workbook()
    sheet = wb.active
    sheet.title = title
    # sheet[‘C3‘] = ‘Hello
    for row, data in enumerate(datas, 1):
        for column, value in enumerate(data, 1):
            sheet.cell(row, column, value)
    print(save_file)
    wb.save(save_file)
    return save_file

# 在使用openpyxl制作excel的过程中突然发现有个错误提示
# raise IllegalCharacterError openpyxl.utils.exceptions.IllegalCharacterError
# 是因为非法字符写入导出出错； 可以通过ILLEGAL_CHARACTERS_RE进行过滤；

def main():
    pass


if __name__ == '__main__':
    main()