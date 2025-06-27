#!/usr/bin/env python
# coding=utf-8

import traceback
import gradio as gr
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy
import tempfile

# 读取Excel文件，筛选A列含关键词的行，E列数值加1，输出新文件。

def read_merged_cells(file_path, sheet_name):
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    merged_ranges = ws.merged_cells.ranges
    merged_info = {}
    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                merged_info[(row, col)] = {
                    'start_row': min_row,
                    'start_col': min_col,
                    'value': ws.cell(row=min_row, column=min_col).value
                }

    df = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

    for row_idx in range(len(df)):
        for col in ['A', 'E', 'F']:
            cell_row = row_idx + 2  # 数据行从Excel第2行开始
            cell_col = column_index_from_string(col)
            if (cell_row, cell_col) in merged_info:
                df.loc[row_idx, col] = merged_info[(cell_row, cell_col)]['value']

    # 处理非合并单元格
    for row_idx in range(len(df)):
        cell_row = row_idx + 2  # 数据行从Excel第2行开始
        cell_col = column_index_from_string('A')
        if (cell_row, cell_col) not in merged_info:
            df.loc[row_idx, 'A'] = ws.cell(row=cell_row, column=cell_col).value

    return df, wb, ws, merged_ranges

def process_excel(file_path, sheet_name, keyword):
    df, wb, ws, merged_ranges = read_merged_cells(file_path, sheet_name)
    if df.empty:
        return None

    mask = df['A'].astype(str).str.contains(keyword, case=False)
    selected_indices = df[mask].index.tolist()

    # 处理E列，增加1
    for idx in selected_indices:
        cell_row = idx + 2  # 数据行从Excel第2行开始
        cell_col = column_index_from_string('E')
        current_value = ws.cell(row=cell_row, column=cell_col).value
        if isinstance(current_value, str):
            try:
                current_value = int(current_value)
                current_value += 1
                ws.cell(row=cell_row, column=cell_col).value = str(current_value)
            except ValueError as e:
                ## 如果无法转换，保持原样
                print('无法转换：', current_value, type(current_value), e)
                print(traceback.format_exc())
        elif isinstance(current_value, int):
            ws.cell(row=cell_row, column=cell_col).value += 1

    selected_rows_1based = set()
    selected_rows_1based.add(1)  # 添加标题行

    for idx in selected_indices:
        row_1based = idx + 2  # 数据行对应Excel行号
        selected_rows_1based.add(row_1based)
        for merged_range in merged_ranges:
            if merged_range.min_row <= row_1based <= merged_range.max_row:
                selected_rows_1based.update(range(merged_range.min_row, merged_range.max_row + 1))

    # 创建新工作簿
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Filtered Results"
    if 'Sheet' in new_wb.sheetnames:
        new_wb.remove(new_wb['Sheet'])

    # 创建行号映射
    row_mapping = {row: idx + 1 for idx, row in enumerate(sorted(selected_rows_1based))}

    # 复制行数据和样式
    for new_row, original_row in enumerate(sorted(selected_rows_1based), start=1):
        for cell in ws[original_row]:
            new_cell = new_ws.cell(row=new_row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)
        # 复制行高
        new_ws.row_dimensions[new_row].height = ws.row_dimensions[original_row].height

    # 处理合并单元格
    for merged_range in merged_ranges:
        merged_rows = set(range(merged_range.min_row, merged_range.max_row + 1))
        if merged_rows & set(selected_rows_1based):
            min_new_row = None
            max_new_row = None
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                if row in row_mapping:
                    mapped_row = row_mapping[row]
                    if min_new_row is None or mapped_row < min_new_row:
                        min_new_row = mapped_row
                    if max_new_row is None or mapped_row > max_new_row:
                        max_new_row = mapped_row
            if min_new_row and max_new_row:
                start_col = get_column_letter(merged_range.min_col)
                end_col = get_column_letter(merged_range.max_col)
                new_ws.merge_cells(f"{start_col}{min_new_row}:{end_col}{max_new_row}")

    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_path = temp_file.name
    new_wb.save(temp_path)
    new_wb.close()

    return temp_path

def update_sheets(file_path):
    if not file_path:
        return []
    try:
        wb = load_workbook(file_path)
        return gr.Dropdown(label="选择Sheet", choices=list(wb.sheetnames))
    except Exception as e:
        print(f"Error: {e}")
        return gr.Dropdown(label="选择Sheet", choices=[])

def main():
    with gr.Blocks() as demo:
        gr.Markdown("Excel 数据筛选工具")

        file_input = gr.File(label="上传Excel文件", type="filepath")
        sheet_select = gr.Dropdown(label="选择Sheet")

        keyword_input = gr.Textbox(label="输入A列筛选关键词")
        process_button = gr.Button("开始筛选")

        output = gr.File(label="下载结果文件")

        file_input.change(
            fn=update_sheets,
            inputs=file_input,
            outputs=sheet_select
        )

        def process_click(file_path, sheet, keyword):
            if not file_path or not sheet or not keyword:
                return None
            return process_excel(file_path, sheet, keyword)

        process_button.click(
            fn=process_click,
            inputs=[file_input, sheet_select, keyword_input],
            outputs=output
        )

    demo.launch(server_name='0.0.0.0', server_port=7860)

if __name__ == "__main__":
    main()