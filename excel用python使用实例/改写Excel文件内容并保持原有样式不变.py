#coding=utf-8

from openpyxl import load_workbook

"""
改写Excel文件的内容，并保证原有的下拉列表，公式，等样式不变。
"""
excel_file = "测试模板文件.xlsx"

inwb = load_workbook(excel_file)

for sheetName in inwb.get_sheet_names():
    sheet = inwb[sheetName]
    # 读取第2行第3列单元格的内容，即 C2 单元格 （行列数不是从0开始，而是从1开始）
    oldstr = sheet.cell(row=2, column=3).value
    print(type(oldstr))
    print(oldstr)
    newstr = '第二行，第3列的新内容；'
    sheet.cell(row=2, column=3).value = newstr 

# 将excel重命名保存
inwb.save("entrust1.xlsx")

